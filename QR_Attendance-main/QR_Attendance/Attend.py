from tkinter import *
import tkinter as tk
from tkinter import ttk, messagebox
import cv2
import pyzbar.pyzbar as pyzbar
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
def load_student_data(filename):
    barcode_to_name = {}
    with open(filename, 'r') as file:
        for line in file:
            parts = line.strip().split(',')  # Assuming each line is "barcode,name"
            if len(parts) == 2:  # Ensure the line is correctly formatted
                barcode, name = parts
                barcode_to_name[barcode] = name
    return barcode_to_name

# Load the barcode to name mapping from an external file
barcode_to_name = load_student_data('students.txt')

def get_or_create_workbook(filename):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Reg No.", "Student Name", "Class & Sec", "Course Code", "In Time"])
    return wb, wb.active

def saveWorkbook(wb, filename):
    wb.save(filename)

def update_sections(event):
    branch_to_sections = {
        "CSE": ["428", "429", "430"],
        "ECE": ["530", "531", "532"],
        "EEE": ["630", "631", "632"],
        "IT": ["730", "731", "732"],
        "MECH": ["830", "831", "832"],
        "ECM": ["930", "931", "932"],
    }
    branch = branch_var.get()
    sections = branch_to_sections.get(branch, [])
    section_combobox['values'] = sections
    if sections:
        section_combobox.current(0)
def update_course_code_sections(event):
    course_code_to_sections = {
        "428": ["A", "B"],  # Example: Course code 428 has sections A and B
        "429": ["A", "B", "C"],  # Course code 429 has sections A, B, C
        "430": ["A", "B","c","D"],  # Example: Course code 428 has sections A and B
        "530": ["A", "B", "C"], 
        "531": ["A", "B"], 
        "532": ["A", "B"], 
        "630": ["A", "B", "C"], 
        "631": ["A", "B", "C"], 
        "632": ["A", "B","c","D"], 
        "730": ["A", "B", "C"], 
        "731": ["A", "B"], 
        "732": ["A", "B", "C"], 
        "830": ["A", "B","c","D"], 
        "831": ["A", "B", "C"], 
        "832": ["A", "B"], 
        "930": ["A", "B", "C"], 
        "931": ["A", "B","c","D"], 
        "932": ["A", "B","c","D"], 

        # Add more mappings as needed
    }
    course_code = section_var.get()
    sections = course_code_to_sections.get(course_code, [])
    sec_combobox['values'] = sections
    if sections:
        sec_combobox.current(0)

def checkk():
    if branch_var.get() and section_var.get() and sec_var.get():
        window.destroy()
    else:
        messagebox.showwarning("Warning", "All fields required!!")

window = tk.Tk()
window.title('Enhanced Attendance System')
window.geometry('900x600')
window.configure(bg="white")  # Set to a lighter background for a cleaner look

# Adjust the title label to have a red background with white text for a strong contrast
title = tk.Label(window, text="Enhanced Attendance System", bd=10, relief=tk.FLAT,
                 font=("times new roman", 40, "bold"), bg="red", fg="white")
title.pack(side=tk.TOP, fill=tk.X)

# Initialize StringVar variables for the comboboxes
branch_var = tk.StringVar()
sec_var = tk.StringVar()
section_var = tk.StringVar()

# Adjust the label and combobox styling for Branch, Section, and Course Code selections
style = ttk.Style()
style.theme_use('default')  # Use the default theme as a base for customization

# Customize the combobox and button styles to fit the red and white theme
style.configure("TCombobox", foreground="black", background="red", fieldbackground="white")
style.configure("TButton", foreground="white", background="red", font=("Times New Roman", 15))

# Branch selection with adjusted label color
ttk.Label(window, text="Branch", background="white", foreground="black",
          font=("Times New Roman", 15)).place(x=100, y=200)
branch_combobox = ttk.Combobox(window, textvariable=branch_var, width=20,
                               font=("times new roman",13), state='readonly', style="TCombobox")
branch_combobox['values'] = ("CSE", "ECE", "EEE", "IT", "MECH", "ECM")
branch_combobox.place(x=250, y=200)
branch_combobox.bind('<<ComboboxSelected>>', update_sections)

# Section selection with updated background and foreground
ttk.Label(window, text="Section", background="white", foreground="black",
          font=("Times New Roman", 15)).place(x=100, y=250)
sec_combobox = ttk.Combobox(window, textvariable=sec_var, width=20,
                            font=("times new roman",13), state='readonly', style="TCombobox")
sec_combobox.place(x=250, y=250)

# Course Code selection with themed styling
ttk.Label(window, text="Course Code", background="white", foreground="black",
          font=("Times New Roman", 15)).place(x=100, y=300)
section_combobox = ttk.Combobox(window, textvariable=section_var, width=20,
                                font=("times new roman",13), state='readonly', style="TCombobox")
section_combobox.place(x=250, y=300)
section_combobox.bind('<<ComboboxSelected>>', update_course_code_sections)

# Submit button with a cohesive theme
exit_button = ttk.Button(window, text="Submit", command=checkk, style="TButton")
exit_button.place(x=300, y=380)

image_path = "cn.png"  # Update this path to your image's path
my_image = PhotoImage(file=image_path)

# Create a label to display the image
image_label = tk.Label(window, image=my_image, bg="white")
image_label.place(x=650, y=150)

window.mainloop()
today = datetime.today().strftime("%Y-%m-%d")
filename = f'{today}_Attendance.xlsx'
wb, ws = get_or_create_workbook(filename)

names = []

def enterData(z):
    if z in names:
        print('Already Present')
    else:
        it = datetime.now()
        names.append(z)
        z_str = ''.join(str(z))  # Assuming 'z' is the barcode
        intime = it.strftime("%H:%M:%S")

        # Lookup the student's name using the barcode
        student_name = barcode_to_name.get(z_str, "Unknown")  # Default to "Unknown" if not found

        # Append both the barcode (z_str) and student name along with other info
        ws.append([z_str, student_name, branch_var.get() + '-' + sec_var.get(), section_var.get(), intime])
        saveWorkbook(wb, filename)

def checkData(data):
    barcode_str = data.decode('utf-8').strip()  # Decode and strip whitespace
    print(f"Scanned Barcode: {barcode_str}")  # Debug print

    if barcode_str in names:
        print('Already Present')
    else:
        print(f'\n{len(names)+1}\npresent...')
        enterData(barcode_str)  # Pass the cleaned barcode data


cap = cv2.VideoCapture(0)

while True:
    _, frame = cap.read()
    decodedObjects = pyzbar.decode(frame)
    for obj in decodedObjects:
        checkData(obj.data)
        time.sleep(1)

    cv2.imshow("Frame", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cv2.destroyAllWindows()
cap.release()
